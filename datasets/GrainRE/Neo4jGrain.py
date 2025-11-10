import re
from py2neo import Graph,Node,Relationship
from py2neo.matching import *
from 储粮害虫数据解析 import *
from bs4 import BeautifulSoup
from 获取害虫名及学名 import getAllNamePro
import os
import openpyxl


# 初始化
graph = Graph("http://localhost:7474/",username="neo4j",password="admin") # 连接Neo4j数据库，输入地址、用户名、密码

#匹配中文字符串
def reMacth(string):
    PATTERN = r'[\u4e00-\u9fa5]+'
    pattern = re.compile(PATTERN)
    p = pattern.findall((string))
    return p

#读取文件中的害虫名字
def getAllName():
    namelist = []
    with open('储粮害虫名字.txt','r',encoding='utf-8') as f:
        for name in f.readlines():
            namelist.append(name[:-1])
    return namelist


#保存害虫名字到Neo4j中
def saveName2Neo4j():
    cnameList = getAllName()
    nodes = NodeMatcher(graph)
    rootNode = nodes.match("储粮害虫", cname='储粮害虫').first()
    if not rootNode :
        rootNode = Node('储粮害虫', cname='储粮害虫')
        graph.create(rootNode)
    for cname in cnameList:
        tempNode = Node('储粮害虫', cname=cname)
        tempRelation = Relationship(rootNode, '包含', tempNode)
        graph.create(tempNode | tempRelation)

#保存害虫名字和学名以及类别到Neo4j中
def saveXName2Neo4j():
    xnameList = getAllNamePro()
    nodes = NodeMatcher(graph)
    rootNode = nodes.match('粮虫',cname='粮虫').first()
    if not rootNode :
        rootNode = Node('粮虫',cname='粮虫')
        graph.create(rootNode)
    for xname in xnameList:
        if xname.name[-1] == '目':
            mname = xname
            muNode = Node('粮虫',cname=mname.name,xname=mname.xname)
            tempRelation = Relationship(rootNode,'包含',muNode)
            graph.create(muNode | tempRelation)
        elif xname.name[-1] == '科':
            kname = xname
            kNode = Node('粮虫',cname=kname.name,xname=kname.xname)
            tempRelation = Relationship(muNode, '包含', kNode)
            graph.create(kNode | tempRelation)
        else:
            tempNode = Node('粮虫',cname=xname.name,xname=xname.xname)
            tempRelation = Relationship(kNode, '包含', tempNode)
            graph.create(tempNode | tempRelation)


#保存害虫基本信息到Neo4j中
def saveBaseInfo2Neo4j():
    filelist = getFilesName()
    nodes = NodeMatcher(graph)
    for name in filelist:
        bs = BeautifulSoup(open(r"baidubaike\\" + name, encoding='utf-8'))
        hcInfo = getBaseInfo(bs)
        cname = hcInfo.get('中文学名') if hcInfo.get('中文学名')  else hcInfo.get('中文名')
        matchNode = nodes.match('储粮害虫',cname = cname).first()
        if not matchNode:
            print('---未找到匹配节点---')
            continue
        for key,value in hcInfo.items():
            if key in ['中文名' ,'中文学名']:
                continue
            tempNode = nodes.match('储粮害虫',cname = value).first()
            if not tempNode:
                tempNode = Node('储粮害虫',cname = value)
                tempRelation = Relationship(matchNode,key,tempNode)
                graph.create(tempNode|tempRelation)
            else:
                tempRelation = Relationship(matchNode, key, tempNode)
                graph.create(tempRelation)

#害虫分布信息到Neo4j中
def saveDistribution():
    wb = openpyxl.load_workbook('data.xlsx')
    sheet = wb.worksheets[0]
    nodes = NodeMatcher(graph)
    for row in sheet.iter_rows():
        cname = row[0].value
        cname = cname.strip()
        distList = reMacth(row[3].value)
        matchNode = nodes.match('粮虫', cname=cname).first()
        if not matchNode:
            print(cname + '---未找到匹配节点---')
            continue
        for dist in distList:
            tempNode = nodes.match('粮虫',cname = dist).first()
            if not tempNode:
                tempNode = Node('粮虫',cname = dist,type = 'distribution')
                tempRelation = Relationship(matchNode,'分布',tempNode)
                graph.create(tempNode|tempRelation)
            else:
                tempRelation = Relationship(matchNode, '分布', tempNode)
                graph.create(tempRelation)

#害虫外观特征信息到Neo4j中
def saveView():
    wb = openpyxl.load_workbook('data.xlsx')
    sheet = wb.worksheets[0]
    nodes = NodeMatcher(graph)
    for row in sheet.iter_rows():
        cname = row[0].value
        cname = cname.strip()
        view = row[2].value
        matchNode = nodes.match('粮虫', cname=cname).first()
        if not matchNode:
            print(cname + '---未找到匹配节点---')
            continue
        tempNode = Node('粮虫',cname = view,type = 'economic')
        tempRelation = Relationship(matchNode,'经济意义',tempNode)
        graph.create(tempNode|tempRelation)



if __name__ == '__main__':
    # saveName2Neo4j()
    # saveBaseInfo2Neo4j()
    # saveXName2Neo4j()
    # saveDistribution()
    saveView()
    print('Done')



