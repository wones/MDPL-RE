import openpyxl
from 获取害虫名及学名 import getAllNamePro

def cnameSaveExcel():
    wb = openpyxl.Workbook()
    sheet = wb.create_sheet(index=0, title='name')
    nameList = getAllNamePro()
    for index, cname in enumerate(nameList):
        sheet.cell(index + 1, 1).value = cname.name
        sheet.cell(index + 1, 2).value = cname.xname

    wb.save('name.xlsx')


if __name__ == '__main__':
    # cnameSaveExcel()
    pass